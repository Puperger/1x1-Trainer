'''Imports'''
import random
import tkinter as tk
from tkinter import ttk
import openpyxl



wb = openpyxl.load_workbook('Score.xlsx')

ws = wb.active

'''Set Variables'''
score=0
total=0
ques = "hello"
rightL=[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
totalL=[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
RTVal=0


'''Define Functions'''
def newcalc():
    global question_label
    global ques
    global res
    global RTVal
    a = random.randint(0,10)#Get Random Numbers between 0 and 10
    b = random.randint(0,10)
    RTVal = b
    astr = str(a) #Make a and b strings
    bstr = str(b)
    xstr = "x"
    ques = f"{a} {xstr} {bstr}"#Combine the Strings
    '''Set the first 2 labels'''
    question_label = ttk.Label(main, text=ques) 
    question_label.grid(column=0, row=0, columnspan=1, sticky=tk.EW)
    scorelabel = ttk.Label(main, text=(score, "/", total))
    scorelabel.grid(column=0, row= 3, sticky=tk.EW)
    #Get Solution
    res= a*b



def ginp():
    global score
    global res
    global total
    global rightL
    global totalL
    global RTVal
    ctrlvar = input_label.get()#get input
    if int(ctrlvar) == int(res):#check if awnser correct
        print("Nice")
        score += 1
        rightL[RTVal] += 1
    else:
        print("The right awnser would've been ",res)
    total += 1
    totalL[RTVal] += 1

    #print(totalL)
    #print(rightL)


    clearentry()#call functions
    newcalc()
    updateXL()

def clearentry():
    input_label.delete(0, 'end')#clear entry label


def updateXL():
    ws['B2'].value = rightL[0]
    ws['B3'].value = rightL[1]
    ws['B4'].value = rightL[2]
    ws['B5'].value = rightL[3]
    ws['B6'].value = rightL[4]
    ws['B7'].value = rightL[5]
    ws['B8'].value = rightL[6]
    ws['B9'].value = rightL[7]
    ws['B10'].value = rightL[8]
    ws['B11'].value = rightL[9]
    ws['B12'].value = rightL[10]
    wb.save('Score.xlsx')

'''Setting the window up'''
main = tk.Tk()
main.geometry("600x300")
main.title('1x1 Trainer')
main.resizable(True, True)

main.columnconfigure(0)
main.columnconfigure(1)
main.rowconfigure(0, weight=1)
main.rowconfigure(1, weight=1)
main.rowconfigure(2, weight=1)
main.rowconfigure(3, weight=1)

newcalc()#start the cycle



'''Define some more labels'''
input_label = ttk.Entry(main)
input_label.grid(column=0, row= 1, columnspan=1, sticky=tk.EW)

controll_button = ttk.Button(main, text="Check", command=ginp)
controll_button.grid(column=0, row=2, columnspan=1, sticky=tk.EW)





main.mainloop()

